import os
import openpyxl as xl

excel_path = os.path.join(os.getcwd(), 'excel')
all_files = [f for f in os.listdir(excel_path) if os.path.isfile(os.path.join(excel_path, f))]

columns = {}

current_new_row = {'global': 1, 'temp': 1}
current_new_column = 1
total_products = 0

wb_final = xl.Workbook()
ws_final = wb_final.active

for file in all_files:

    file_path = os.path.join(excel_path, file)
    wb = xl.load_workbook(file_path)
    ws = wb.active
    total_products += ws.max_row - 1

    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        current_new_row['temp'] = current_new_row['global']
        if header not in columns.keys():
            columns.update({header: current_new_column})
            ws_final.cell(1, current_new_column).value = header
            current_new_row['temp'] += 1
            for row in range(2, ws.max_row + 1):
                print(f'file: {file}, row: {row}, col: {col}')
                ws_final.cell(current_new_row['temp'], current_new_column).value = ws.cell(row, col).value
                current_new_row['temp'] += 1
            current_new_column += 1
        else:
            for row in range(2, ws.max_row + 1):
                print(f'file: {file}, row: {row}, col: {col}')
                ws_final.cell(current_new_row['temp'], columns[header]).value = ws.cell(row, col).value
                current_new_row['temp'] += 1
    current_new_row['global'] = current_new_row['temp']

wb_final.save('final.xlsx')
# print('total products:', total_products)
