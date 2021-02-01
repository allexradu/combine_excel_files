import os
import openpyxl as xl

excel_path = os.path.join(os.getcwd(), 'excel')
all_files = [f for f in os.listdir(excel_path) if os.path.isfile(os.path.join(excel_path, f))]
illegal_characters = ["' '", "''", "'"]
value = ''


for file in all_files:
    file_path = os.path.join(excel_path, file)
    wb = xl.load_workbook(file_path)
    ws = wb.active

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            for character in illegal_characters:
                print(f'file: {file}, row: {row}, col: {col}')
                cell_value = ws.cell(row, col).value
                if cell_value is not None:
                    if character in cell_value:
                        ws.cell(row, col).value = ws.cell(row, col).value.replace(character, '')
    wb.save(file_path)
